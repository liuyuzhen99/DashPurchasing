import openpyxl
import pandas as pd
from traversal_file_test import traversalDir_SecDir
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

file_path_start = r"C:\Users\eqhw2il\OneDrive - 奥迪一汽新能源汽车有限公司\Shared Documents - NEV Co. QPNi\A6L e-tron (AU511_4CE_BL)"


# file_path = r"C:\Users\eqhw2il\OneDrive - 奥迪一汽新能源汽车有限公司\Shared Documents - NEV Co. QPNi\A6L e-tron (AU511_4CE_BL)\B10_Black Panel linked to Q6L\02_Maturity Levels\QPNi_EXTERNAL AU511 PID-COVER-25042023.xlsx"


def calculate_status(file_path, ml):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[ml]
    df = pd.read_excel(file_path, sheet_name=ml)
    line_num = len(df)
    R = 0
    G = 0
    Y = 0
    for i in range(1, line_num + 1):
        # print("--------", sheet.cell(row=i, column=12).value)
        if sheet.cell(row=i, column=12).value == 1:
            G += 1
        elif sheet.cell(row=i, column=12).value == 2:
            Y += 1
        elif sheet.cell(row=i, column=12).value == 3:
            R += 1
    if R != 0:
        return 'red'
    elif Y != 0:
        return 'yellow'
    elif G != 0:
        return 'green'
    else:
        return 'gray'


def ml_sum_status(file_path_start, ml):
    lists = traversalDir_SecDir(file_path_start)
    ml_list = [0, 0, 0]
    for li in lists:
        li_ml_status = calculate_status(li, ml)
        if li_ml_status == 'red':
            ml_list[0] += 1
        elif li_ml_status == 'yellow':
            ml_list[1] += 1
        elif li_ml_status == 'green':
            ml_list[2] += 1
    return ml_list

# wb = openpyxl.load_workbook(file_path, data_only=True)
# sheet = wb['ml2']
# df = pd.read_excel(file_path, sheet_name='ml2')
# line_num = len(df)
# print(df.columns)
#
# R = 0
# G = 0
# Y = 0
#
# for i in range(1, line_num + 1):
#     print("--------", sheet.cell(row=i, column=12).value)
#     if sheet.cell(row=i, column=12).value == 1:
#         G += 1
#     elif sheet.cell(row=i, column=12).value == 2:
#         Y += 1
#     elif sheet.cell(row=i, column=12).value == 3:
#         R += 1
#
# print("R:", R)
# print("G:", G)
# print("Y:", Y)
#
# # print(sheet.cell(row=2, column=12).value)
# print(sheet.dimensions)
