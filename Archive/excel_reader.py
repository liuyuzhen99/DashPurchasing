import glob
import os
import openpyxl
import pandas as pd
import warnings

class Excel_Reader:
    def __init__(self, path_start=r"C:\Users\eqhw2il\OneDrive - 奥迪一汽新能源汽车有限公司\Shared Documents - NEV Co. QPNi\A6L e-tron (AU511_4CE_BL)") -> None:
        self.path_start = path_start
    
    def traversalDir_SecDir(self):
        lists = []
        if os.path.exists(self.path_start):
            level1_files = glob.glob(self.path_start + '\\*')
            for level2_file in level1_files:
                if os.path.isdir(level2_file):
                    level2_files = glob.glob(level2_file + '\\*')
                    for level3_file in level2_files:
                        if os.path.isdir(level3_file):
                            h = os.path.split(level3_file)
                            if "02_Maturity Level" in h[1]:
                                level3_files = glob.glob(level3_file + '\\*.xlsx')
                                for level4_file in level3_files:
                                    lists.append(level4_file)
        return lists

    def calculate_status(self, path, ml):
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb[ml]
        df = pd.read_excel(path, sheet_name=ml)
        line_num = len(df)
        R = 0
        G = 0
        Y = 0
        for i in range(1, line_num + 1):
            if sheet.cell(row=i, column=12).value == 1:
                G += 1
            elif sheet.cell(row=i, column=12).value == 2:
                Y += 1
            elif sheet.cell(row=i, column=12).value == 3:
                R += 1
        if R != 0:
            return 'red'
        elif Y != 0:
            return 'yellow'
        elif G != 0:
            return 'green'
        else:
            return 'gray'


    def ml_sum_status(self, ml):
        lists = self.traversalDir_SecDir()
        ml_list = [0, 0, 0]
        for li in lists:
            li_ml_status = self.calculate_status(li, ml)
            if li_ml_status == 'red':
                ml_list[0] += 1
            elif li_ml_status == 'yellow':
                ml_list[1] += 1
            elif li_ml_status == 'green':
                ml_list[2] += 1
        return ml_list
    
    def compare_colors(self, color1, color2):
        color_order = {'gray': 0, 'green': 1, 'yellow': 2, 'red': 3}

        return color_order[color1] - color_order[color2]
    
    def ml2_7_status_per_file(self, file_path):
        status = 'gray'
        for i in range(2,8):
            temp = self.calculate_status(file_path,f'ml{i}')
            temp_result = self.compare_colors(temp, status)
            if temp_result >= 0:
                status = temp
        return status
